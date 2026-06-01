"""Shared parsing helpers for RBI State Finances xlsx tables."""

from __future__ import annotations

import re
from datetime import date
from decimal import Decimal, InvalidOperation
from typing import Iterable

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

# RBI uses '–' (en dash) and other Unicode quirks in cell labels.
_DASH_RE = re.compile(r"[–—−]")
_FY_PATTERN = re.compile(r"(\d{4})\s*-\s*(\d{2,4})")
_BASIS_TOKENS = (
    ("accounts", "audited_actual"),
    ("(actuals)", "audited_actual"),
    ("actuals", "audited_actual"),
    ("re", "revised_estimate"),
    ("(re)", "revised_estimate"),
    ("revised", "revised_estimate"),
    ("be", "budget_estimate"),
    ("(be)", "budget_estimate"),
    ("budget", "budget_estimate"),
    ("provisional", "monthly_actual_provisional"),
)


def open_first_sheet(path: str) -> Worksheet:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    return wb[wb.sheetnames[0]]


def open_workbook(path: str) -> openpyxl.Workbook:
    return openpyxl.load_workbook(path, data_only=True, read_only=True)


def normalize_text(value: object) -> str:
    """Strip Unicode dashes, collapse whitespace, return lowercase text."""
    if value is None:
        return ""
    text = _DASH_RE.sub("-", str(value))
    return " ".join(text.split()).strip()


def to_decimal(value: object) -> Decimal | None:
    """Coerce a cell value to Decimal. Treat '–', empty, '..', 'N.A.' as None."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        try:
            return Decimal(repr(value))
        except (InvalidOperation, ValueError):
            return None
    text = normalize_text(value)
    if not text:
        return None
    if text in {"-", "..", "...", "n.a.", "n/a", "na", "nil"}:
        return None
    cleaned = text.replace(",", "").replace(" ", "")
    if cleaned in {"-", ""}:
        return None
    try:
        return Decimal(cleaned)
    except InvalidOperation:
        return None


def fiscal_year_from_label(label: object) -> str | None:
    """Extract '2025-26' from inputs like '2025-26 (RE)', 2025, '2025–26'."""
    if label is None:
        return None
    text = normalize_text(label)
    if not text:
        return None
    m = _FY_PATTERN.search(text)
    if m:
        start = m.group(1)
        end_raw = m.group(2)
        end = end_raw if len(end_raw) == 2 else end_raw[-2:]
        return f"{start}-{end}"
    if text.isdigit() and len(text) == 4:
        # Bare calendar year used as 'as-at-end-March YYYY' — treat as the FY
        # that ends in YYYY. So 2014 → 2013-14.
        year = int(text)
        return f"{year - 1}-{str(year)[-2:]}"
    return None


def basis_from_label(label: object, *, default: str = "audited_actual") -> str:
    """Detect basis tag from a column label like '2024-25 (RE)'."""
    text = normalize_text(label).lower()
    for token, basis in _BASIS_TOKENS:
        if token in text:
            return basis
    return default


def fiscal_period(fiscal_year: str) -> tuple[date, date]:
    """'2025-26' → (2025-04-01, 2026-03-31)."""
    start_year = int(fiscal_year.split("-")[0])
    return date(start_year, 4, 1), date(start_year + 1, 3, 31)


def as_at_march_period(year_label: str | int) -> tuple[date, date, str]:
    """A bare year column in Statement 19 means as-at-end-March that year.

    Returns (period_start, period_end, fiscal_year_label).
    """
    year = int(year_label)
    fy = f"{year - 1}-{str(year)[-2:]}"
    return date(year - 1, 4, 1), date(year, 3, 31), fy


def iter_rows_with_index(ws: Worksheet) -> Iterable[tuple[int, tuple]]:
    """Yields (1-based row_number, row tuple) for each row in the sheet."""
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        yield i, row


_HEADER_ANCHOR_TOKENS = ("state/ut", "states/ut", "state/uts", "states/uts")


def is_state_header_cell(cell: object) -> bool:
    """True only for short header-row labels like 'State/UT' (not the long table title)."""
    text = normalize_text(cell).lower()
    if not text or len(text) > 30:
        return False
    text_compact = text.replace(" ", "")
    return any(token.replace(" ", "") in text_compact for token in _HEADER_ANCHOR_TOKENS)


def find_header_row(ws: Worksheet, *, anchor: str = "state/ut") -> int:  # noqa: ARG001
    """Find the 1-based row that holds 'State/UT'-style column labels.

    The ``anchor`` arg is kept for API compatibility but ignored; matching is
    done via ``is_state_header_cell`` so we never lock onto title rows that
    happen to start with the word 'Statement'.
    """
    for row_idx, row in iter_rows_with_index(ws):
        if row_idx > 14:
            break
        for cell in row:
            if is_state_header_cell(cell):
                return row_idx
    raise ValueError(f"Could not find State/UT header row in sheet '{ws.title}'")


def parse_year_columns_table(
    path,
    *,
    metric_code: str,
    metric_name: str,
    metric_group: str,
    unit: str,
    unit_scale: str,
):
    """Generic parser for RBI tables shaped as:

      header row : 'State/UT', 2008, 2009, ..., 2026[, '2025-26 (RE)', '2026-27 (BE)']
      next row   : 1, 2, 3, ...  (column-number row, skipped)
      data rows  : '<n>. <State>', value, value, ...

    Used by Statement 19, 20, 28. Yields FiscalMetricRow per (state, year) cell.
    """
    from worker.state_finances_xlsx.records import FiscalMetricRow, XlsxProvenance
    from worker.state_finances_xlsx.state_codes import normalize_state

    ws = open_first_sheet(str(path))
    header_row_idx = find_header_row(ws)
    rows = list(ws.iter_rows(values_only=True))
    header_row = rows[header_row_idx - 1]

    year_columns: list[tuple[int, int, str]] = []
    for col_idx, cell in enumerate(header_row, start=1):
        if col_idx <= 2:
            continue
        text = normalize_text(cell)
        if not text:
            continue
        lower = text.lower()
        if "(be)" in lower or " be" in lower or "budget estimate" in lower:
            basis = "budget_estimate"
        elif "(re)" in lower or " re" in lower or "revised estimate" in lower or "revised" in lower:
            basis = "revised_estimate"
        elif "(p)" in lower or "provisional" in lower:
            basis = "monthly_actual_provisional"
        elif "accounts" in lower or "(actuals)" in lower:
            basis = "audited_actual"
        else:
            basis = "audited_actual"
        if text.isdigit() and len(text) == 4:
            year_columns.append((col_idx, int(text), basis))
            continue
        fy = fiscal_year_from_label(text)
        if fy:
            year = int(fy.split("-")[0]) + 1
            year_columns.append((col_idx, year, basis))
            continue
        # Bare year with basis suffix, e.g. '2025 (RE)'
        for token in text.split():
            tok = token.strip("(),")
            if tok.isdigit() and len(tok) == 4:
                year_columns.append((col_idx, int(tok), basis))
                break

    sheet_name = ws.title
    for row_idx, row in enumerate(rows[header_row_idx:], start=header_row_idx + 1):
        if not row:
            continue
        if is_column_number_row(row):
            continue
        state_cell = row[1] if len(row) > 1 else None
        state_code = normalize_state(state_cell)
        if state_code is None:
            continue
        for col_idx, year, basis in year_columns:
            if col_idx - 1 >= len(row):
                continue
            value = to_decimal(row[col_idx - 1])
            if value is None:
                continue
            period_start, period_end, fiscal_year = as_at_march_period(year)
            yield FiscalMetricRow(
                state_code=state_code,
                metric_code=metric_code,
                metric_name=metric_name,
                metric_group=metric_group,
                basis_tag=basis,
                fiscal_year=fiscal_year,
                period_start=period_start,
                period_end=period_end,
                value=value,
                unit=unit,
                unit_scale=unit_scale,
                provenance=XlsxProvenance(
                    sheet_name=sheet_name,
                    row_number=row_idx,
                    column_index=col_idx,
                    column_label=str(year),
                    row_label=str(state_cell),
                ),
            )


def is_column_number_row(row: tuple) -> bool:
    """True if the row looks like RBI's '1, 2, 3, ...' column-number row."""
    found_first = False
    seen = 0
    for cell in row:
        if cell is None:
            continue
        text = normalize_text(cell)
        if not text:
            continue
        if not found_first:
            if text == "1":
                found_first = True
                seen = 1
                continue
            return False
        # any further cells should be small ints
        try:
            int_val = int(text)
        except ValueError:
            return False
        seen += 1
        if int_val > 30:  # column count never exceeds the table width
            return False
    return found_first and seen >= 2
